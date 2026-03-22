"""
Document Processing Utilities for Automotive Market Data

Handles extraction and processing of:
- PDFs: Automotive policies, regulations, market research, strategic documents
- Excel files: Market statistics, OEM performance data, EV adoption metrics, sales trends
- CSV files: Tabular data for forecasting
"""

import os
import logging
from pathlib import Path
import PyPDF2
import openpyxl
import pandas as pd
from werkzeug.utils import secure_filename

# Allowed file extensions
ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'xls', 'csv'}
logger = logging.getLogger(__name__)
BACKEND_DIR = Path(__file__).resolve().parents[2]
UPLOADS_DIR = BACKEND_DIR / 'uploads'


class DocumentProcessingError(Exception):
    """Raised when a document cannot be parsed into usable text."""

def build_unique_upload_path(filename, upload_dir=None):
    """
    Build a unique upload path so new uploads never overwrite an existing file.

    Args:
        filename (str): The original filename.
        upload_dir (str): Directory where files are stored.

    Returns:
        tuple: (stored_filename, absolute_path)
    """
    target_dir = Path(upload_dir) if upload_dir else UPLOADS_DIR
    target_dir.mkdir(parents=True, exist_ok=True)

    base_name, extension = os.path.splitext(filename)
    stored_filename = filename
    file_path = str(target_dir / stored_filename)
    counter = 1

    while os.path.exists(file_path):
        stored_filename = f"{base_name}_{counter}{extension}"
        file_path = str(target_dir / stored_filename)
        counter += 1

    return stored_filename, file_path

def allowed_file(filename):
    """
    Check if the file extension is allowed.

    Args:
        filename (str): The filename to check.

    Returns:
        bool: True if allowed, False otherwise.
    """
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_text_from_pdf(file_path):
    """
    Extract text from a PDF file.

    Args:
        file_path (str): Path to the PDF file.

    Returns:
        str: Extracted text.
    """
    errors = []

    # Loader 1: PyPDF2 (baseline)
    try:
        parts = []
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                page_text = page.extract_text() or ""
                if page_text.strip():
                    parts.append(page_text)
        text = "\n".join(parts).strip()
        if text:
            return text
        errors.append("PyPDF2 extracted no text")
    except Exception as exc:
        errors.append(f"PyPDF2 failed: {exc}")

    # Loader 2: pdfplumber (optional dependency)
    try:
        import pdfplumber  # type: ignore

        parts = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                if page_text.strip():
                    parts.append(page_text)
        text = "\n".join(parts).strip()
        if text:
            return text
        errors.append("pdfplumber extracted no text")
    except ImportError:
        errors.append("pdfplumber not installed")
    except Exception as exc:
        errors.append(f"pdfplumber failed: {exc}")

    # Loader 3: PyMuPDF/fitz (optional dependency)
    try:
        import fitz  # type: ignore

        parts = []
        doc = fitz.open(file_path)
        try:
            for page in doc:
                page_text = page.get_text("text") or ""
                if page_text.strip():
                    parts.append(page_text)
        finally:
            doc.close()

        text = "\n".join(parts).strip()
        if text:
            return text
        errors.append("PyMuPDF extracted no text")
    except ImportError:
        errors.append("PyMuPDF not installed")
    except Exception as exc:
        errors.append(f"PyMuPDF failed: {exc}")

    raise DocumentProcessingError(
        "PDF parsing failed or produced no text. "
        "The file may be scanned/image-only, empty, or corrupted. "
        f"Details: {' | '.join(errors)}"
    )

def extract_text_from_excel(file_path):
    """
    Extract text from an Excel file.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        str: Extracted text.
    """
    text = ""
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join([str(cell) for cell in row if cell is not None])
                if row_text.strip():
                    text += row_text + "\n"
    except Exception as exc:
        raise DocumentProcessingError(f"Excel parse error: {exc}") from exc

    if not text.strip():
        raise DocumentProcessingError("Excel file contains no extractable text")

    return text

def _read_csv_robust(file_path):
    """
    Try to read a CSV with multiple encodings and auto-detected delimiter.
    Returns a DataFrame on success, raises on total failure.
    """
    import csv as _csv

    # Encoding candidates in priority order
    encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252', 'utf-16']

    for enc in encodings:
        try:
            # Sniff delimiter from first 4 KB
            with open(file_path, 'r', encoding=enc, errors='replace') as fh:
                sample = fh.read(4096)
            try:
                dialect = _csv.Sniffer().sniff(sample, delimiters=',;\t|')
                sep = dialect.delimiter
            except _csv.Error:
                sep = ','  # fall back to comma

            df = pd.read_csv(file_path, sep=sep, encoding=enc, on_bad_lines='skip')
            if df.empty or df.shape[1] == 0:
                continue
            return df
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception:
            continue

    # Last resort: read with latin-1 (never fails) and comma separator
    return pd.read_csv(file_path, sep=',', encoding='latin-1', on_bad_lines='skip')


def extract_text_from_csv(file_path):
    """
    Extract text from a CSV file.

    Args:
        file_path (str): Path to the CSV file.

    Returns:
        str: Extracted text from CSV rows.
    """
    try:
        df = _read_csv_robust(file_path)
        text = df.to_string(index=False)
    except Exception as exc:
        raise DocumentProcessingError(f"CSV parse error: {exc}") from exc

    if not text.strip():
        raise DocumentProcessingError("CSV file contains no extractable text")

    return text

def process_document(file):
    """
    Process an uploaded file and extract text.

    Args:
        file: The uploaded file object.

    Returns:
        tuple: (extracted_text, filename)
    """
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

    filename = secure_filename(file.filename)
    if not filename:
        raise DocumentProcessingError("Invalid filename")

    stored_filename, file_path = build_unique_upload_path(filename)

    file.save(file_path)

    ext = os.path.splitext(stored_filename)[1].lower()
    if ext == '.pdf':
        text = extract_text_from_pdf(file_path)
    elif ext in ('.xlsx', '.xls'):
        text = extract_text_from_excel(file_path)
    elif ext == '.csv':
        text = extract_text_from_csv(file_path)
    else:
        raise DocumentProcessingError(f"Unsupported file type: {ext}")

    if not text or not text.strip():
        raise DocumentProcessingError("Extracted text is empty")

    logger.info("Processed %s successfully (%d chars)", stored_filename, len(text))
    return text, stored_filename